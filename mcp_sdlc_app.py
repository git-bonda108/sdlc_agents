import streamlit as st
import os
import json
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Union, get_args, get_origin
from io import BytesIO
import asyncio
from dotenv import load_dotenv
from pydantic import BaseModel, ValidationError

try:
    from autogen import ConversableAgent
except ImportError:
    st.error("AutoGen not installed. Run: pip install pyautogen")
    st.stop()

try:
    import docx
    import openpyxl
except ImportError:
    st.error("Document processing libraries missing. Run: pip install python-docx openpyxl")
    st.stop()

try:
    from streamlit_mermaid import st_mermaid
    HAS_MERMAID = True
except ImportError:
    HAS_MERMAID = False

load_dotenv()

ROOT = Path(__file__).parent
MEMORY_DIR = Path(os.environ.get("MEMORY_DIR", ROOT / "memory"))
MEMORY_DIR.mkdir(exist_ok=True)
DEFAULT_JSON = MEMORY_DIR / "run_memory.json"
PHASES = ["Requirements", "Design", "Build", "IntegrationTest", "UAT", "Cutover"]

AGENT_DESCRIPTIONS = {
    "Manager": "🎯 Orchestrates the SDLC, reviews all outputs, and assigns the best model for each phase.",
    "Requirements": "📋 Analyzes your requirements and documents, builds user stories, and creates a traceability matrix.",
    "Design": "🏗️ Designs the system architecture, components, and data models for your project.",
    "Build": "⚙️ Plans the implementation, repository structure, and code modules for development.",
    "IntegrationTest": "🔧 Designs and automates integration tests to ensure all parts work together.",
    "UAT": "👥 Crafts user acceptance tests to validate business value and user experience.",
    "Cutover": "🚀 Plans deployment, rollback, and communication for a smooth go-live."
}

MODEL_POOLS = {
    "Manager": ["gpt-4o"],
    "Requirements": ["gpt-4o", "claude-3-opus-20240229"],
    "Design": ["gpt-4o", "deepseek-coder", "claude-3-opus-20240229"],
    "Build": ["deepseek-coder", "gpt-4o", "grok-1"],
    "IntegrationTest": ["gpt-4o", "deepseek-coder", "grok-1"],
    "UAT": ["claude-3-opus-20240229", "gpt-4o", "grok-1"],
    "Cutover": ["gpt-4o", "grok-1", "deepseek-coder"]
}

STATUS_MESSAGES = {
    "Requirements": "📋 Requirements Agent is analyzing your requirements and will build user stories. The Manager will review the output. Hang tight!",
    "Design": "🏗️ Design Agent is creating the system architecture and technical specs. The Manager will review the design.",
    "Build": "⚙️ Build Agent is planning the implementation and code modules. The Manager will review the build plan.",
    "IntegrationTest": "🔧 Integration Test Agent is designing and automating integration tests. The Manager will review the test plan.",
    "UAT": "👥 UAT Agent is crafting user acceptance tests for business validation. The Manager will review the UAT plan.",
    "Cutover": "🚀 Cutover Agent is planning deployment and go-live. The Manager will review the cutover plan."
}

def show_sidebar():
    st.sidebar.header("🤖 SDLC Agentic AI Team")
    for phase in ["Manager"] + PHASES:
        st.sidebar.markdown(f"**{phase} Agent**")
        st.sidebar.markdown(AGENT_DESCRIPTIONS[phase])
        if phase in MODEL_POOLS:
            st.sidebar.markdown(
                f"<span style='color:gray;font-size:0.9em'>Models: {', '.join(MODEL_POOLS[phase])}</span>",
                unsafe_allow_html=True
            )
        st.sidebar.markdown("---")

class MCPMessage(BaseModel):
    run_id: str
    phase: str
    role: str
    agent: str
    timestamp: str
    content: Dict[str, Any]
    metadata: Optional[Dict[str, Any]] = None

class RunMemory:
    def __init__(self, path: Path = DEFAULT_JSON):
        self.path = path
        self._load()

    def _load(self):
        if self.path.exists():
            try:
                self.messages: List[Dict[str, Any]] = json.loads(self.path.read_text())
            except json.JSONDecodeError:
                self.messages = []
        else:
            self.messages = []

    def append(self, msg: MCPMessage):
        self.messages.append(msg.model_dump())
        self._flush()

    def _flush(self):
        try:
            self.path.write_text(json.dumps(self.messages, indent=2, ensure_ascii=False))
        except Exception:
            pass

    def get_phase_output(self, phase: str) -> Optional[Dict[str, Any]]:
        for m in reversed(self.messages):
            if m["phase"] == phase and m["role"] == "assistant":
                return m["content"]
        return None

class RequirementsSchema(BaseModel):
    functional_requirements: List[Dict[str, Any]]
    non_functional_requirements: List[Dict[str, Any]]
    project_type: str
    user_stories: List[Dict[str, Any]]
    requirements_traceability_matrix: Dict[str, Any]
    acceptance_criteria: List[Dict[str, Any]]
    stakeholders: List[Dict[str, Any]]
    constraints: List[Dict[str, Any]]

SCHEMA_MAP = {
    "Requirements": RequirementsSchema,
    # For other phases, do not enforce schema
}

AGENT_PROMPTS = {
    "Manager": (
        "You are the Project Manager Agent overseeing a comprehensive SDLC process. "
        "Your role is to evaluate the quality and completeness of outputs from each phase. "
        "If any phase output is insufficient, empty, contains stringified data, or lacks required fields, "
        "provide specific feedback and recommend trying a different model. "
        "You have access to multiple AI models and can orchestrate which model should handle each phase."
    ),
    "Requirements": (
        "You are an expert Requirements Analyst. "
        "Analyze the project goal and any uploaded documents to produce comprehensive requirements. "
        "MANDATORY: Fill ALL fields with detailed, actionable content. Never leave fields empty. "
        "If you do not know a value, fill with a plausible placeholder. "
        "Output ONLY valid JSON, no explanations, no markdown, no extra text. "
        "EXAMPLE OUTPUT:\n"
        "{\n"
        "  \"functional_requirements\": [{\"id\": \"FR-001\", \"description\": \"User login via SSO\", \"priority\": \"High\", \"dependencies\": []}],\n"
        "  \"non_functional_requirements\": [{\"id\": \"NFR-001\", \"description\": \"System must handle 1000 concurrent users\", \"priority\": \"High\"}],\n"
        "  \"project_type\": \"Enterprise Web Application\",\n"
        "  \"user_stories\": [{\"id\": \"US-001\", \"title\": \"User Login\", \"description\": \"As a new employee, I want to log in using SSO\", \"acceptance_criteria\": [\"Login redirects to SSO provider\"]}],\n"
        "  \"requirements_traceability_matrix\": {\"FR-001\": \"US-001\", \"NFR-001\": \"US-001\"},\n"
        "  \"acceptance_criteria\": [{\"requirement_id\": \"FR-001\", \"criteria\": \"User successfully logs in via SSO\"}],\n"
        "  \"stakeholders\": [{\"name\": \"HR Department\", \"role\": \"Business Owner\", \"contactInformation\": {\"email\": \"hr@company.com\"}}],\n"
        "  \"constraints\": [{\"id\": \"C-001\", \"description\": \"Must use Azure AD\", \"impact\": \"Limits authentication options\"}]\n"
        "}"
    ),
    "Design": (
        "You are a Senior Software Architect. Based on the requirements, create a comprehensive system design. "
        "MANDATORY: Use well-formatted Markdown, bulleted lists, tables, and diagrams for clarity. "
        "If you do not know a value, fill with a plausible placeholder. "
        "Output should be readable and business-friendly."
    ),
    "Build": (
        "You are a Lead Developer responsible for implementation planning. "
        "Create a detailed build plan with repository structure, implementation steps, and code module descriptions. "
        "MANDATORY: Use well-formatted Markdown, bulleted lists, and tables for clarity. "
        "If you do not know a value, fill with a plausible placeholder. "
        "Output should be readable and business-friendly."
    ),
    "IntegrationTest": (
        "You are a Senior QA Engineer specializing in integration testing. "
        "Design comprehensive integration tests with detailed test plans and automation scripts. "
        "MANDATORY: Use well-formatted Markdown, bulleted lists, and tables for clarity. "
        "If you do not know a value, fill with a plausible placeholder. "
        "Output should be readable and business-friendly."
    ),
    "UAT": (
        "You are a Business Analyst expert in User Acceptance Testing. "
        "Create comprehensive UAT materials that business users can execute. "
        "MANDATORY: Use well-formatted Markdown, bulleted lists, and tables for clarity. "
        "If you do not know a value, fill with a plausible placeholder. "
        "Output should be readable and business-friendly."
    ),
    "Cutover": (
        "You are a DevOps Engineer and Deployment Specialist. "
        "Create a comprehensive cutover plan with detailed deployment procedures. "
        "MANDATORY: Use well-formatted Markdown, bulleted lists, and tables for clarity. "
        "If you do not know a value, fill with a plausible placeholder. "
        "Output should be readable and business-friendly."
    )
}

def get_llm_config(model: str) -> Dict[str, Any]:
    base_config = {
        "temperature": 0.1,
        "timeout": int(os.environ.get("OPENAI_TIMEOUT", 600)),
        "max_tokens": 4000,
    }
    if model.startswith("grok"):
        api_key = os.environ.get("GROK_API_KEY")
        if not api_key:
            st.warning(f"Grok API key not found, falling back to GPT-4o")
            return get_llm_config("gpt-4o")
        return {
            **base_config,
            "model": model,
            "api_key": api_key,
            "api_type": "openai",
            "base_url": "https://api.x.ai/v1"
        }
    elif model.startswith("deepseek"):
        api_key = os.environ.get("DEEPSEEK_API_KEY")
        if not api_key:
            st.warning(f"DeepSeek API key not found, falling back to GPT-4o")
            return get_llm_config("gpt-4o")
        return {
            **base_config,
            "model": model,
            "api_key": api_key,
            "api_type": "openai",
            "base_url": "https://api.deepseek.com/v1"
        }
    elif model.startswith("claude"):
        api_key = os.environ.get("ANTHROPIC_API_KEY")
        if not api_key:
            st.warning("Anthropic API key not found, falling back to OpenAI GPT-4o")
            return get_llm_config("gpt-4o")
        return {
            **base_config,
            "model": model,
            "api_key": api_key,
            "api_type": "anthropic"
        }
    else:
        api_key = os.environ.get("OPENAI_API_KEY")
        if not api_key:
            st.error("OpenAI API key is required")
            st.stop()
        return {
            **base_config,
            "model": model,
            "api_key": api_key,
        }

def make_agent(name: str, system_prompt: str, model: str = "gpt-4o") -> ConversableAgent:
    config = get_llm_config(model)
    return ConversableAgent(
        name=name,
        llm_config=config,
        system_message=system_prompt,
        human_input_mode="NEVER",
        code_execution_config=False,
        max_consecutive_auto_reply=1
    )

def build_agents() -> Dict[str, ConversableAgent]:
    agents = {}
    for phase in ["Manager"] + PHASES:
        model_pool = MODEL_POOLS.get(phase, ["gpt-4o"])
        primary_model = model_pool[0]
        try:
            agents[phase] = make_agent(f"{phase}Agent", AGENT_PROMPTS.get(phase, f"You are the {phase} Agent."), primary_model)
        except Exception:
            agents[phase] = make_agent(f"{phase}Agent", AGENT_PROMPTS.get(phase, f"You are the {phase} Agent."), "gpt-4o")
    return agents

def parse_docx(file: BytesIO) -> str:
    try:
        doc = docx.Document(file)
        text_content = []
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                text_content.append(paragraph.text.strip())
        return "\n".join(text_content)
    except Exception as e:
        st.error(f"Error parsing DOCX file: {e}")
        return ""

def parse_xlsx(file: BytesIO) -> str:
    try:
        wb = openpyxl.load_workbook(file)
        text_content = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text_content.append(f"Sheet: {sheet_name}")
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    text_content.append(row_text)
        return "\n".join(text_content)
    except Exception as e:
        st.error(f"Error parsing XLSX file: {e}")
        return ""

def is_output_insufficient(phase: str, parsed_content: Dict[str, Any]) -> bool:
    if not isinstance(parsed_content, dict):
        return True
    if "error" in parsed_content:
        return True
    schema_cls = SCHEMA_MAP.get(phase)
    if not schema_cls:
        return False
    required_fields = list(schema_cls.model_fields.keys())
    for field in required_fields:
        value = parsed_content.get(field)
        if value is None or value == "":
            return True
        if isinstance(value, (list, dict)) and not value:
            return True
        if isinstance(value, str) and value.lower() in ["todo", "tbd", "placeholder", "example"]:
            return True
    return False

def coerce_to_schema(schema_cls, data):
    if data is None:
        origin = get_origin(schema_cls)
        if origin in (list, List):
            return []
        if origin in (dict, Dict):
            return {}
        if schema_cls == str:
            return ""
        return data
    if isinstance(schema_cls, type) and issubclass(schema_cls, BaseModel):
        coerced = {}
        for field, field_info in schema_cls.model_fields.items():
            expected_type = field_info.annotation
            value = data.get(field) if isinstance(data, dict) else None
            coerced[field] = coerce_to_schema(expected_type, value)
        return coerced
    origin = get_origin(schema_cls)
    args = get_args(schema_cls)
    if origin in (list, List):
        if isinstance(data, list):
            return [coerce_to_schema(args[0] if args else str, v) for v in data]
        elif isinstance(data, str):
            return [data]
        elif isinstance(data, dict):
            return [str(v) for v in data.values()]
        else:
            return [str(data)]
    if origin in (dict, Dict):
        if isinstance(data, dict):
            if len(args) >= 2:
                return {str(k): coerce_to_schema(args[1], v) for k, v in data.items()}
            else:
                return {str(k): str(v) for k, v in data.items()}
        elif isinstance(data, list):
            result = {}
            for idx, item in enumerate(data):
                if isinstance(item, dict):
                    key = item.get("name") or item.get("component") or f"Item {idx+1}"
                    val = item.get("description") or item.get("desc") or str(item)
                    result[key] = val
                elif isinstance(item, str):
                    result[f"Item {idx+1}"] = item
            return result
        elif isinstance(data, str):
            return {"value": data}
        else:
            return {"value": str(data)}
    if hasattr(schema_cls, "__origin__") and schema_cls.__origin__ is Union:
        for arg in schema_cls.__args__:
            if arg == type(None):
                continue
            try:
                return coerce_to_schema(arg, data)
            except Exception:
                continue
        return str(data) if data is not None else ""
    return data

async def run_sdlc_phase_orchestrated(
    phase: str,
    context_msg: str,
    memory: RunMemory,
    run_id: str,
    manager_agent: ConversableAgent
) -> Dict[str, Any]:
    model_pool = MODEL_POOLS.get(phase, ["gpt-4o"])
    raw_outputs = []
    for model_idx, model in enumerate(model_pool):
        try:
            st.info(f"🤖 {phase} Phase: Trying {model} (attempt {model_idx + 1}/{len(model_pool)})")
            agent = make_agent(f"{phase}Agent", AGENT_PROMPTS.get(phase, f"You are the {phase} Agent."), model)
            response = agent.generate_reply(
                messages=[{"role": "user", "content": context_msg}]
            )
            if isinstance(response, dict) and "content" in response:
                content = response["content"]
            elif isinstance(response, str):
                content = response
            else:
                content = str(response)
            raw_outputs.append({"model": model, "raw": content})

            # For Requirements, enforce schema
            if phase == "Requirements":
                try:
                    parsed_content = json.loads(content)
                except json.JSONDecodeError:
                    import re
                    json_match = re.search(r'\{.*\}', content, re.DOTALL)
                    if json_match:
                        try:
                            parsed_content = json.loads(json_match.group())
                        except:
                            parsed_content = {"output": content, "raw_response": True}
                    else:
                        parsed_content = {"output": content, "raw_response": True}
                schema_cls = SCHEMA_MAP.get(phase)
                if schema_cls and not parsed_content.get("raw_response"):
                    coerced = coerce_to_schema(schema_cls, parsed_content)
                    try:
                        validated = schema_cls.model_validate(coerced)
                        parsed_content = validated.model_dump()
                    except ValidationError as e:
                        st.warning(f"Schema validation failed for {phase} with {model}: {e}")
                        if model_idx < len(model_pool) - 1:
                            continue
                        else:
                            parsed_content = coerced
                if is_output_insufficient(phase, parsed_content):
                    manager_feedback = (
                        f"Output from {model} for {phase} phase is insufficient. "
                        f"Missing or empty required fields detected. "
                        f"Trying next model in pool: {model_pool[model_idx + 1] if model_idx + 1 < len(model_pool) else 'None (last attempt)'}"
                    )
                    memory.append(MCPMessage(
                        run_id=run_id,
                        phase=phase,
                        role="manager",
                        agent="ManagerAgent",
                        timestamp=datetime.now(timezone.utc).isoformat(),
                        content={"manager_feedback": manager_feedback, "model_used": model},
                        metadata={"insufficient_output": True, "attempt": model_idx + 1}
                    ))
                    st.warning(f"⚠️ {model} output insufficient for {phase}. Trying next model...")
                    if model_idx < len(model_pool) - 1:
                        continue
                    else:
                        st.error(f"❌ All models failed for {phase} phase")
                        return {
                            "error": f"All models in pool failed for {phase} phase",
                            "models_tried": model_pool,
                            "phase": phase,
                            "raw_outputs": raw_outputs
                        }
                else:
                    st.success(f"✅ {phase} Phase completed successfully with {model}")
                    memory.append(MCPMessage(
                        run_id=run_id,
                        phase=phase,
                        role="assistant",
                        agent=f"{phase}Agent",
                        timestamp=datetime.now(timezone.utc).isoformat(),
                        content=parsed_content,
                        metadata={
                            "model_used": model,
                            "attempt": model_idx + 1,
                            "total_attempts": len(model_pool),
                            "tokens_used": len(str(parsed_content))
                        }
                    ))
                    return parsed_content
            else:
                # For all other phases, accept any output and display as Markdown
                st.success(f"✅ {phase} Phase completed successfully with {model}")
                memory.append(MCPMessage(
                    run_id=run_id,
                    phase=phase,
                    role="assistant",
                    agent=f"{phase}Agent",
                    timestamp=datetime.now(timezone.utc).isoformat(),
                    content={"output": content},
                    metadata={
                        "model_used": model,
                        "attempt": model_idx + 1,
                        "total_attempts": len(model_pool),
                        "tokens_used": len(str(content))
                    }
                ))
                return {"output": content}
        except Exception as e:
            st.warning(f"⚠️ Error with {model} for {phase}: {str(e)}")
            if model_idx < len(model_pool) - 1:
                await asyncio.sleep(2)
                continue
            else:
                error_response = {
                    "error": f"All models failed for {phase} phase",
                    "last_error": str(e),
                    "models_tried": model_pool,
                    "phase": phase,
                    "raw_outputs": raw_outputs
                }
                memory.append(MCPMessage(
                    run_id=run_id,
                    phase=phase,
                    role="assistant",
                    agent=f"{phase}Agent",
                    timestamp=datetime.now(timezone.utc).isoformat(),
                    content=error_response,
                    metadata={"failed": True, "models_tried": model_pool}
                ))
                return error_response

def render_table(data, title=None):
    if not data:
        return
    if title:
        st.markdown(f"### {title}")
    if isinstance(data, dict):
        if all(isinstance(v, (str, int, float, bool, type(None))) for v in data.values()):
            st.table([{"Key": k, "Value": str(v)} for k, v in data.items()])
        else:
            for k, v in data.items():
                st.markdown(f"**{k}:**")
                if isinstance(v, (dict, list)):
                    render_table(v)
                else:
                    st.write(v)
                st.markdown("---")
    elif isinstance(data, list):
        if all(isinstance(item, dict) for item in data):
            if data:
                st.table(data)
        else:
            for i, item in enumerate(data):
                st.markdown(f"**{i+1}.** {item}")
    else:
        st.write(data)

def render_requirements_output(req_output):
    render_table(req_output.get("functional_requirements"), "Functional Requirements")
    render_table(req_output.get("non_functional_requirements"), "Non-Functional Requirements")
    st.markdown("### Project Type")
    st.write(req_output.get("project_type", ""))
    render_table(req_output.get("user_stories"), "User Stories")
    if req_output.get("requirements_traceability_matrix"):
        st.markdown("### Requirements Traceability Matrix")
        matrix = req_output["requirements_traceability_matrix"]
        if isinstance(matrix, dict):
            st.table([{"Requirement": k, "Mapping": str(v)} for k, v in matrix.items()])
        else:
            st.write(matrix)
    render_table(req_output.get("acceptance_criteria"), "Acceptance Criteria")
    render_table(req_output.get("stakeholders"), "Stakeholders")
    render_table(req_output.get("constraints"), "Constraints")

def render_phase_output(phase, output):
    if not output:
        st.info("No output for this phase.")
        return
    if phase == "Requirements":
        render_requirements_output(output)
    else:
        # If output is a dict with a single "output" key, treat as Markdown
        if isinstance(output, dict) and "output" in output:
            val = output["output"]
            try:
                # Try to parse as JSON for pretty display
                parsed = json.loads(val)
                st.json(parsed)
            except Exception:
                st.markdown(val)
        elif isinstance(output, str):
            try:
                parsed = json.loads(output)
                st.json(parsed)
            except Exception:
                st.markdown(output)
        else:
            st.write(output)

async def run_sdlc(
    top_level_goal: str,
    memory: RunMemory,
    requirements_file: Optional[Any] = None
):
    run_id = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    agents = build_agents()
    manager_agent = agents["Manager"]
    progress_bar = st.progress(0)
    status_placeholder = st.empty()
    phase_outputs = {}

    # Requirements Phase
    current_phase = "Requirements"
    progress_bar.progress(1/len(PHASES))
    with status_placeholder.container():
        st.info(STATUS_MESSAGES[current_phase])
    requirements_context = f"TOP LEVEL GOAL:\n{top_level_goal}\n\n"
    if requirements_file:
        with st.spinner("📄 Parsing uploaded requirements document..."):
            file_content = requirements_file.read()
            if requirements_file.name.endswith(".docx"):
                requirements_text = parse_docx(BytesIO(file_content))
            elif requirements_file.name.endswith(".xlsx"):
                requirements_text = parse_xlsx(BytesIO(file_content))
            else:
                st.error("❌ Unsupported file type. Please upload a DOCX or XLSX file.")
                return
            requirements_context += f"REQUIREMENTS DOCUMENT ({requirements_file.name}):\n{requirements_text}\n\n"
            requirements_context += "Please analyze this document and extract comprehensive requirements, generate user stories, and create a requirements traceability matrix."
    else:
        requirements_context += "No requirements document provided. Please generate comprehensive requirements based on the top-level goal."
    phase_output = await run_sdlc_phase_orchestrated(
        current_phase,
        requirements_context,
        memory,
        run_id,
        manager_agent
    )
    if "error" in phase_output:
        st.error(f"❌ {current_phase} phase failed: {phase_output['error']}")
        if "raw_outputs" in phase_output:
            st.warning("Raw LLM outputs for debugging:")
            for entry in phase_output["raw_outputs"]:
                st.code(f"Model: {entry['model']}\n\n{entry['raw']}", language="json")
        return
    phase_outputs[current_phase] = phase_output
    with status_placeholder.container():
        st.success(f"✅ {current_phase} Complete: Requirements analyzed and user stories generated successfully!")
    time.sleep(1)

    # Subsequent Phases
    for i, phase in enumerate(PHASES[1:], 1):
        progress_bar.progress((i + 1) / len(PHASES))
        with status_placeholder.container():
            st.info(STATUS_MESSAGES[phase])
        context_msg = f"TOP LEVEL GOAL:\n{top_level_goal}\n\nPREVIOUS PHASE OUTPUTS:\n"
        for prev_phase, prev_output in phase_outputs.items():
            context_msg += f"\n{prev_phase.upper()}:\n{json.dumps(prev_output, indent=2)}\n"
        phase_output = await run_sdlc_phase_orchestrated(
            phase,
            context_msg,
            memory,
            run_id,
            manager_agent
        )
        if "error" in phase_output:
            st.error(f"❌ {phase} phase failed: {phase_output['error']}")
            if "raw_outputs" in phase_output:
                st.warning("Raw LLM outputs for debugging:")
                for entry in phase_output["raw_outputs"]:
                    st.code(f"Model: {entry['model']}\n\n{entry['raw']}", language="json")
            return
        phase_outputs[phase] = phase_output
        with status_placeholder.container():
            st.success(f"✅ {phase} Complete: {phase} phase completed successfully!")
        time.sleep(1)

    # Completion
    progress_bar.progress(1.0)
    with status_placeholder.container():
        st.success("🎉 SDLC COMPLETE! All phases have been successfully completed. Your project is ready for deployment!")
    st.balloons()
    st.success(f"✅ SDLC run {run_id} completed successfully!")

    st.header("📊 SDLC Results")
    for phase in PHASES:
        phase_output = memory.get_phase_output(phase)
        if phase_output:
            with st.expander(f"✅ {phase} Phase Output", expanded=False):
                render_phase_output(phase, phase_output)
        else:
            with st.expander(f"⏳ {phase} Phase (Not Started)", expanded=False):
                st.info("This phase has not been completed yet.")

def main():
    st.set_page_config(page_title="AI SDLC Platform", page_icon="🚀", layout="wide")
    st.title("🚀 AI-Powered SDLC Platform")
    st.markdown("**Multi-Model Orchestration • Agentic AI • Best-in-Class Output Quality**")
    show_sidebar()
    memory = RunMemory()
    st.subheader("📋 Project Configuration")
    goal = st.text_area(
        "Top-level project goal",
        height=120,
        placeholder="e.g., Build an employee onboarding platform that integrates with Azure AD, automates IT account provisioning, provides self-service capabilities, and ensures compliance with GDPR and SOC 2..."
    )
    requirements_file = st.file_uploader(
        "📄 Upload Requirements Document (Optional)",
        type=["docx", "xlsx"],
        help="Upload a DOCX or XLSX file containing detailed project requirements"
    )
    st.subheader("🎮 Controls")
    run_btn = st.button("▶️ Start Multi-Model SDLC Run", type="primary")
    if run_btn and goal.strip():
        with st.container():
            st.subheader("🚀 Multi-Model SDLC Execution in Progress")
            asyncio.run(run_sdlc(goal.strip(), memory, requirements_file))
    elif run_btn and not goal.strip():
        st.error("❌ Please enter a project goal before starting the SDLC run.")

if __name__ == "__main__":
    main()